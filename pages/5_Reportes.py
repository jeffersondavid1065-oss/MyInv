import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime
from sqlalchemy import text
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import obtener_conexion
from queries import (
    obtener_ventas_periodo,
    obtener_top_productos,
    obtener_metricas_mes,
    obtener_facturas_periodo,
)
from utils import aplicar_estilos, verificar_auth
from tz_utils import hoy_bogota, ahora_bogota
from alegra_utils import facturar_venta

st.set_page_config(page_title="Reportes", layout="wide")
aplicar_estilos()
user_id, nombre_negocio = verificar_auth()

engine = obtener_conexion()

def formato_cop(numero):
    return f"${float(numero):,.0f}".replace(",", ".")


def generar_excel_ventas(df_ventas, fecha_ini, fecha_fin, ingresos, costos, margen, nombre_negocio):
    """Genera reporte Excel de ventas del período."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    descuentos = float(df_ventas['descuento'].sum()) if not df_ventas.empty else 0

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = f"REPORTE DE VENTAS — {nombre_negocio}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{fecha_ini.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    # Resumen financiero
    ws.append(["RESUMEN FINANCIERO", "", "", "", "", "", "", ""])
    ws["A4"].font = Font(bold=True, size=11, color="1F4E78")

    ws.append(["Ingresos Totales", f"${ingresos:,.0f}".replace(",", ".")])
    ws.append(["Costo de Ventas", f"${costos:,.0f}".replace(",", ".")])
    ws.append(["Descuentos Totales", f"${descuentos:,.0f}".replace(",", ".")])
    ws.append(["Margen Bruto", f"${margen:,.0f}".replace(",", ".")])
    ws.append([])

    # Headers de tabla
    headers = ["ID Venta", "Fecha", "Cliente", "Descuento ($)", "Total ($)", "Tipo Pago", "Estado", "Cambio ($)"]
    ws.append(headers)
    header_row = ws.max_row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    if not df_ventas.empty:
        for _, row in df_ventas.iterrows():
            ws.append([
                row.get('id', ''),
                str(row.get('fecha', '')),
                row.get('cliente', 'Venta directa'),
                float(row.get('descuento', 0)),
                float(row.get('total', 0)),
                row.get('tipo_pago', ''),
                row.get('estado', ''),
                float(row.get('cambio', 0)),
            ])
            for col_num in range(1, 9):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = border
                if col_num in [4, 5, 8]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

    # Total
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=3).value = "TOTAL"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=3).fill = total_fill
    ws.cell(row=total_row, column=4).value = descuentos
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5).value = ingresos
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


st.title("Reportes")
st.markdown(f"Análisis de ventas para: **{nombre_negocio}**")
st.markdown("---")

tab_dia, tab_periodo, tab_productos, tab_cartera, tab_facturas = st.tabs([
    "Ventas del Día",
    "Reporte por Período",
    "Productos Más Vendidos",
    "Cartera y Créditos",
    "Facturación Electrónica"
])

# ==========================================
# TAB 1: VENTAS DEL DÍA
# ==========================================
with tab_dia:
    st.subheader(f"Ventas de Hoy — {hoy_bogota().strftime('%d/%m/%Y')}")

    hoy = hoy_bogota()
    df_hoy = obtener_ventas_periodo(user_id, hoy, hoy)

    if not df_hoy.empty:
        # Métricas
        total_hoy = df_hoy['total'].sum()
        efectivo_hoy = df_hoy['monto_efectivo'].sum()
        transfer_hoy = df_hoy['monto_transferencia'].sum()
        descuentos_hoy = df_hoy['descuento'].sum()
        cant_ventas = len(df_hoy)
        ticket_prom = total_hoy / cant_ventas if cant_ventas > 0 else 0

        col1, col2, col3, col4, col5, col6 = st.columns(6)
        col1.metric("Ventas", cant_ventas)
        col2.metric("Total Ingresos", formato_cop(total_hoy))
        col3.metric("Caja Efectivo", formato_cop(efectivo_hoy))
        col4.metric("Transferencias", formato_cop(transfer_hoy))
        col5.metric("Ticket Promedio", formato_cop(ticket_prom))
        col6.metric("Descuentos", formato_cop(descuentos_hoy))

        st.markdown("---")

        # Tabla detallada
        st.dataframe(
            df_hoy[['id', 'fecha', 'cliente', 'subtotal', 'descuento', 'total', 'tipo_pago', 'estado']].rename(columns={
                'id': 'ID', 'fecha': 'Hora', 'cliente': 'Cliente',
                'subtotal': 'Subtotal ($)', 'descuento': 'Descuento ($)',
                'total': 'Total ($)', 'tipo_pago': 'Pago', 'estado': 'Estado'
            }),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Subtotal ($)": st.column_config.NumberColumn(format="$%,d"),
                "Descuento ($)": st.column_config.NumberColumn(format="$%,d"),
                "Total ($)": st.column_config.NumberColumn(format="$%,d"),
            }
        )

        # Ver detalle de una venta
        st.markdown("---")
        st.markdown("**Ver ítems de una venta:**")
        dict_ventas_hoy = {
            f"Venta #{r['id']} — {formato_cop(r['total'])}": r['id']
            for _, r in df_hoy.iterrows()
        }
        venta_sel_hoy = st.selectbox("Selecciona la venta", options=list(dict_ventas_hoy.keys()), key="venta_hoy")
        venta_id_hoy = dict_ventas_hoy[venta_sel_hoy]

        with engine.connect() as conn:
            df_items_hoy = pd.read_sql_query(text("""
                SELECT nombre_producto, cantidad, precio_unitario, descuento, subtotal
                FROM Detalles_Venta WHERE venta_id = :vid
            """), con=conn, params={"vid": venta_id_hoy})

        if not df_items_hoy.empty:
            st.dataframe(
                df_items_hoy.rename(columns={
                    'nombre_producto': 'Producto', 'cantidad': 'Cant.',
                    'precio_unitario': 'Precio Unit.', 'descuento': 'Descuento', 'subtotal': 'Subtotal'
                }),
                use_container_width=True, hide_index=True,
                column_config={
                    "Precio Unit.": st.column_config.NumberColumn(format="$%,d"),
                    "Descuento": st.column_config.NumberColumn(format="$%,d"),
                    "Subtotal": st.column_config.NumberColumn(format="$%,d"),
                }
            )
    else:
        st.info("No hay ventas registradas hoy todavía.")

# ==========================================
# TAB 2: REPORTE POR PERÍODO
# ==========================================
with tab_periodo:
    st.subheader("Reporte de Ventas por Período")

    col_f1, col_f2 = st.columns(2)
    with col_f1:
        hoy = hoy_bogota()
        hace_30 = hoy - timedelta(days=30)
        fechas_rango = st.date_input("Rango de fechas", [hace_30, hoy], key="fechas_periodo")
    with col_f2:
        hoy_dt = ahora_bogota()
        mes_sel = st.selectbox("O selecciona un mes", range(1, 13),
                               index=hoy_dt.month - 1,
                               format_func=lambda m: ["Enero", "Febrero", "Marzo", "Abril",
                                                       "Mayo", "Junio", "Julio", "Agosto",
                                                       "Septiembre", "Octubre", "Noviembre",
                                                       "Diciembre"][m - 1])
        año_sel = st.number_input("Año", value=hoy_dt.year, min_value=2020)

    usar_mes = st.checkbox("Usar selección de mes completo", value=False)

    if usar_mes:
        fecha_ini_rep = date(año_sel, mes_sel, 1)
        if mes_sel == 12:
            fecha_fin_rep = date(año_sel + 1, 1, 1) - timedelta(days=1)
        else:
            fecha_fin_rep = date(año_sel, mes_sel + 1, 1) - timedelta(days=1)
    else:
        if len(fechas_rango) == 2:
            fecha_ini_rep, fecha_fin_rep = fechas_rango
        else:
            fecha_ini_rep = fecha_fin_rep = hoy_bogota()

    df_periodo = obtener_ventas_periodo(user_id, fecha_ini_rep, fecha_fin_rep)

    # Métricas del período
    ingresos_rep = float(df_periodo['total'].sum()) if not df_periodo.empty else 0
    efectivo_rep = float(df_periodo['monto_efectivo'].sum()) if not df_periodo.empty else 0
    transfer_rep = float(df_periodo['monto_transferencia'].sum()) if not df_periodo.empty else 0
    descuentos_rep = float(df_periodo['descuento'].sum()) if not df_periodo.empty else 0

    # Costos del período
    with engine.connect() as conn:
        costos_rep = conn.execute(text("""
            SELECT COALESCE(SUM(dv.costo_unitario * dv.cantidad), 0)
            FROM Detalles_Venta dv
            JOIN Ventas v ON dv.venta_id = v.id
            WHERE v.usuario_id = :uid
            AND DATE(v.fecha) >= :f_ini AND DATE(v.fecha) <= :f_fin
            AND v.estado != 'Anulada'
        """), {
            "uid": user_id,
            "f_ini": fecha_ini_rep.strftime('%Y-%m-%d'),
            "f_fin": fecha_fin_rep.strftime('%Y-%m-%d')
        }).scalar()
        costos_rep = float(costos_rep) if costos_rep else 0

    margen_rep = ingresos_rep - costos_rep

    col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
    col_r1.metric("Ingresos", formato_cop(ingresos_rep))
    col_r2.metric("Efectivo", formato_cop(efectivo_rep))
    col_r3.metric("Transferencias", formato_cop(transfer_rep))
    col_r4.metric("Descuentos", formato_cop(descuentos_rep))
    col_r5.metric("Margen Bruto", formato_cop(margen_rep),
                  delta_color="inverse" if margen_rep < 0 else "normal")

    st.markdown("---")

    if not df_periodo.empty:
        st.dataframe(
            df_periodo[['id', 'fecha', 'cliente', 'subtotal', 'descuento', 'total', 'tipo_pago', 'estado']].rename(columns={
                'id': 'ID', 'fecha': 'Fecha', 'cliente': 'Cliente',
                'subtotal': 'Subtotal ($)', 'descuento': 'Descuento ($)',
                'total': 'Total ($)', 'tipo_pago': 'Pago', 'estado': 'Estado'
            }),
            use_container_width=True, hide_index=True,
            column_config={
                "Subtotal ($)": st.column_config.NumberColumn(format="$%,d"),
                "Descuento ($)": st.column_config.NumberColumn(format="$%,d"),
                "Total ($)": st.column_config.NumberColumn(format="$%,d"),
            }
        )

        # Descargar Excel
        st.markdown("---")
        excel_buf = generar_excel_ventas(
            df_periodo, fecha_ini_rep, fecha_fin_rep,
            ingresos_rep, costos_rep, margen_rep, nombre_negocio
        )
        st.download_button(
            label="Descargar Reporte en Excel (para DIAN)",
            data=excel_buf,
            file_name=f"Ventas_{fecha_ini_rep}_{fecha_fin_rep}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("No hay ventas en este período.")

# ==========================================
# TAB 3: PRODUCTOS MÁS VENDIDOS
# ==========================================
with tab_productos:
    st.subheader("Top Productos Más Vendidos")

    col_tp1, col_tp2 = st.columns(2)
    with col_tp1:
        hoy = hoy_bogota()
        hace_30 = hoy - timedelta(days=30)
        fechas_top = st.date_input("Período", [hace_30, hoy], key="fechas_top")
    with col_tp2:
        limite_top = st.number_input("Cantidad de productos", min_value=5, max_value=50, value=10)

    if len(fechas_top) == 2:
        f_ini_top, f_fin_top = fechas_top
        df_top = obtener_top_productos(user_id, f_ini_top, f_fin_top, int(limite_top))

        if not df_top.empty:
            col_tp_a, col_tp_b = st.columns([2, 1])

            with col_tp_a:
                st.bar_chart(
                    df_top.set_index('nombre_producto')['unidades_vendidas'],
                    height=350
                )

            with col_tp_b:
                st.markdown("**Ranking:**")
                for i, (_, row) in enumerate(df_top.iterrows(), 1):
                    st.write(f"**{i}.** {row['nombre_producto']}")
                    st.caption(f"{int(row['unidades_vendidas'])} unidades — {formato_cop(row['total_vendido'])}")

            st.markdown("---")
            st.dataframe(
                df_top.rename(columns={
                    'nombre_producto': 'Producto',
                    'unidades_vendidas': 'Unidades Vendidas',
                    'total_vendido': 'Total Vendido ($)'
                }),
                use_container_width=True, hide_index=True,
                column_config={
                    "Total Vendido ($)": st.column_config.NumberColumn(format="$%,d"),
                }
            )
        else:
            st.info("No hay ventas en este período.")

# ==========================================
# TAB 4: CARTERA Y CRÉDITOS
# ==========================================
with tab_cartera:
    st.subheader("Resumen de Cartera")

    with engine.connect() as conn:
        df_cartera = pd.read_sql_query(text("""
            SELECT cl.nombre as cliente, cl.telefono,
                   COUNT(cr.id) as creditos_activos,
                   SUM(cr.total) as total_prestado,
                   SUM(cr.saldo_pendiente) as saldo_pendiente,
                   MAX(cr.fecha_limite) as proxima_fecha,
                   SUM(CASE WHEN v.factura_estado = 'emitida' THEN 1 ELSE 0 END) as creditos_facturados
            FROM Creditos cr
            JOIN Clientes cl ON cr.cliente_id = cl.id
            JOIN Ventas v ON cr.venta_id = v.id
            WHERE cr.usuario_id = :uid AND cr.estado = 'Activo'
            GROUP BY cl.id, cl.nombre, cl.telefono
            ORDER BY saldo_pendiente DESC
        """), con=conn, params={"uid": user_id})

    if not df_cartera.empty:
        total_cartera = df_cartera['saldo_pendiente'].sum()
        st.metric("Total en Cartera", formato_cop(total_cartera))

        st.dataframe(
            df_cartera.rename(columns={
                'cliente': 'Cliente', 'telefono': 'Teléfono',
                'creditos_activos': 'Créditos', 'total_prestado': 'Total Prestado ($)',
                'saldo_pendiente': 'Saldo Pendiente ($)', 'proxima_fecha': 'Próximo Vence',
                'creditos_facturados': 'Facturados'
            }),
            use_container_width=True, hide_index=True,
            column_config={
                "Total Prestado ($)": st.column_config.NumberColumn(format="$%,d"),
                "Saldo Pendiente ($)": st.column_config.NumberColumn(format="$%,d"),
            }
        )

        # Descargar cartera en Excel
        wb_c = Workbook()
        ws_c = wb_c.active
        ws_c.title = "Cartera"
        ws_c.append(["REPORTE DE CARTERA — " + nombre_negocio])
        ws_c.append([f"Generado: {hoy_bogota().strftime('%d/%m/%Y')}"])
        ws_c.append([])
        ws_c.append(["Cliente", "Teléfono", "Créditos Activos", "Total Prestado", "Saldo Pendiente", "Próximo Vence"])
        for _, row in df_cartera.iterrows():
            ws_c.append([
                row['cliente'], row['telefono'],
                int(row['creditos_activos']),
                float(row['total_prestado']),
                float(row['saldo_pendiente']),
                str(row['proxima_fecha'])
            ])
        ws_c.append([])
        ws_c.append(["", "", "", "TOTAL CARTERA", float(total_cartera), ""])

        buf_c = BytesIO()
        wb_c.save(buf_c)
        buf_c.seek(0)

        st.download_button(
            label="Descargar Cartera en Excel",
            data=buf_c,
            file_name=f"Cartera_{hoy_bogota()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.success("No hay créditos activos. ¡Cartera limpia!")

# ==========================================
# TAB 5: FACTURACIÓN ELECTRÓNICA
# ==========================================
with tab_facturas:
    st.subheader("Historial de Facturación Electrónica")

    col_ff1, col_ff2 = st.columns(2)
    with col_ff1:
        hoy_fe = hoy_bogota()
        hace_30_fe = hoy_fe - timedelta(days=30)
        fechas_fe = st.date_input("Período", [hace_30_fe, hoy_fe], key="fechas_facturas")

    if len(fechas_fe) == 2:
        f_ini_fe, f_fin_fe = fechas_fe
        df_facturas = obtener_facturas_periodo(user_id, f_ini_fe, f_fin_fe)

        if not df_facturas.empty:
            emitidas = (df_facturas['factura_estado'] == 'emitida').sum()
            con_error = (df_facturas['factura_estado'] == 'error').sum()
            anuladas_nc = (df_facturas['factura_estado'] == 'anulada').sum()
            sin_facturar = df_facturas['factura_estado'].isna().sum()

            col_e1, col_e2, col_e3, col_e4, col_e5 = st.columns(5)
            col_e1.metric("Ventas del período", len(df_facturas))
            col_e2.metric("Facturadas", int(emitidas))
            col_e3.metric("Con error", int(con_error),
                          delta="Revisar" if con_error > 0 else None,
                          delta_color="inverse")
            col_e4.metric("Anuladas (N.C.)", int(anuladas_nc))
            col_e5.metric("Sin facturar", int(sin_facturar))

            st.markdown("---")

            col_s1, col_s2, col_s3 = st.columns(3)
            with col_s1:
                busq_id = st.text_input("Buscar por ID de venta", key="busq_factura_id")
            with col_s2:
                busq_num_factura = st.text_input(
                    "Buscar por N° de factura (prefijo o número)", key="busq_factura_numero"
                )
            with col_s3:
                filtro_estado = st.selectbox(
                    "Filtrar por estado",
                    ["Todas", "Facturadas", "Con error", "Anuladas (N.C.)", "Sin facturar"],
                    key="filtro_estado_factura"
                )

            col_s4, col_s5, col_s6 = st.columns(3)
            with col_s4:
                busq_nit = st.text_input("Buscar por NIT/documento del cliente", key="busq_factura_nit")
            with col_s5:
                busq_nombre = st.text_input("Buscar por nombre del cliente", key="busq_factura_nombre")
            with col_s6:
                busq_fecha = st.date_input("Buscar por fecha exacta", value=None, key="busq_factura_fecha")

            df_mostrar_fe = df_facturas.copy()
            if filtro_estado == "Facturadas":
                df_mostrar_fe = df_mostrar_fe[df_mostrar_fe['factura_estado'] == 'emitida']
            elif filtro_estado == "Con error":
                df_mostrar_fe = df_mostrar_fe[df_mostrar_fe['factura_estado'] == 'error']
            elif filtro_estado == "Anuladas (N.C.)":
                df_mostrar_fe = df_mostrar_fe[df_mostrar_fe['factura_estado'] == 'anulada']
            elif filtro_estado == "Sin facturar":
                df_mostrar_fe = df_mostrar_fe[df_mostrar_fe['factura_estado'].isna()]

            df_mostrar_fe = df_mostrar_fe.copy()
            df_mostrar_fe['numero_factura_texto'] = (
                df_mostrar_fe['factura_prefijo'].fillna('').astype(str)
                + df_mostrar_fe['factura_numero'].fillna('').astype(str)
            )

            if busq_id.strip():
                df_mostrar_fe = df_mostrar_fe[
                    df_mostrar_fe['id'].astype(str).str.contains(busq_id.strip(), case=False, na=False)
                ]
            if busq_num_factura.strip():
                df_mostrar_fe = df_mostrar_fe[
                    df_mostrar_fe['numero_factura_texto'].str.contains(busq_num_factura.strip(), case=False, na=False)
                ]
            if busq_nit.strip():
                df_mostrar_fe = df_mostrar_fe[
                    df_mostrar_fe['cliente_documento'].fillna('').astype(str).str.contains(busq_nit.strip(), case=False, na=False)
                ]
            if busq_nombre.strip():
                df_mostrar_fe = df_mostrar_fe[
                    df_mostrar_fe['cliente'].str.contains(busq_nombre.strip(), case=False, na=False)
                ]
            if busq_fecha:
                df_mostrar_fe = df_mostrar_fe[
                    pd.to_datetime(df_mostrar_fe['fecha']).dt.date == busq_fecha
                ]

            df_mostrar_fe['estado_texto'] = df_mostrar_fe['factura_estado'].fillna('Sin facturar').replace({
                'emitida': 'Emitida', 'error': 'Error', 'anulada': 'Anulada (N.C.)'
            })

            st.dataframe(
                df_mostrar_fe[['id', 'fecha', 'cliente', 'cliente_documento', 'total', 'estado_texto',
                               'numero_factura_texto', 'factura_cufe', 'nota_credito_alegra_id']].rename(columns={
                    'id': 'Venta #', 'fecha': 'Fecha', 'cliente': 'Cliente',
                    'cliente_documento': 'NIT/Documento',
                    'total': 'Total ($)', 'estado_texto': 'Estado',
                    'numero_factura_texto': 'N° Factura', 'factura_cufe': 'CUFE',
                    'nota_credito_alegra_id': 'ID Nota Crédito'
                }),
                use_container_width=True, hide_index=True,
                column_config={
                    "Total ($)": st.column_config.NumberColumn(format="$%,d"),
                }
            )

            dict_pdfs = {}
            for _, r in df_mostrar_fe[df_mostrar_fe['factura_pdf_url'].notna()].iterrows():
                etiqueta = f"Venta #{r['id']} — {r['cliente']} — {formato_cop(r['total'])} — Factura"
                dict_pdfs[etiqueta] = r['factura_pdf_url']
            for _, r in df_mostrar_fe[df_mostrar_fe['nota_credito_pdf_url'].notna()].iterrows():
                etiqueta = f"Venta #{r['id']} — {r['cliente']} — {formato_cop(r['total'])} — Nota Crédito"
                dict_pdfs[etiqueta] = r['nota_credito_pdf_url']

            if dict_pdfs:
                st.markdown("---")
                st.markdown("**Descargar factura o nota crédito:**")
                pdf_sel = st.selectbox("Selecciona el documento", options=list(dict_pdfs.keys()), key="pdf_factura_sel")
                st.link_button("Abrir PDF", dict_pdfs[pdf_sel], use_container_width=True)

            facturas_con_error = df_facturas[df_facturas['factura_estado'] == 'error']
            if not facturas_con_error.empty:
                st.markdown("---")
                st.markdown("**Reintentar facturas con error:**")
                dict_reintento = {
                    f"Venta #{r['id']} — {r['cliente']} — {formato_cop(r['total'])}": r['id']
                    for _, r in facturas_con_error.iterrows()
                }
                reintento_sel = st.selectbox("Selecciona la venta", options=list(dict_reintento.keys()), key="reintento_factura_sel")
                if st.button("Reintentar emisión", use_container_width=True):
                    with st.spinner("Reintentando..."):
                        ok_r, msg_r = facturar_venta(user_id, dict_reintento[reintento_sel])
                    if ok_r:
                        st.success(msg_r)
                        obtener_facturas_periodo.clear()
                        st.rerun()
                    else:
                        st.error(msg_r)
        else:
            st.info("No hay ventas en este período.")
