import os
import streamlit as st
import pandas as pd
from datetime import timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from queries import (
    CATEGORIAS_GASTO,
    crear_gasto,
    actualizar_gasto,
    eliminar_gasto,
    obtener_gastos_periodo,
    obtener_gasto,
    obtener_gastos_por_categoria,
)
from utils import aplicar_estilos, verificar_auth, bloquear_si_cajero
from tz_utils import hoy_bogota, ahora_bogota_naive

st.set_page_config(page_title="Gastos", layout="wide")
aplicar_estilos()
user_id, nombre_negocio = verificar_auth()
bloquear_si_cajero()

COMPROBANTES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "comprobantes_gastos")
os.makedirs(COMPROBANTES_DIR, exist_ok=True)

TIPOS_PAGO_GASTO = ["Efectivo", "Transferencia"]


def formato_cop(numero):
    return f"${float(numero):,.0f}".replace(",", ".")


def guardar_comprobante(archivo, uid):
    """Guarda el comprobante subido en disco y retorna su ruta. None si no hay archivo."""
    if not archivo:
        return None
    ext = archivo.name.split(".")[-1].lower()
    nombre_archivo = f"gasto_{uid}_{int(ahora_bogota_naive().timestamp())}.{ext}"
    ruta = os.path.join(COMPROBANTES_DIR, nombre_archivo)
    with open(ruta, "wb") as f:
        f.write(archivo.getbuffer())
    return ruta


def generar_excel_gastos(df_gastos, fecha_ini, fecha_fin, nombre_negocio):
    """Genera reporte Excel de gastos del período, mismo estilo que Reportes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = f"REPORTE DE GASTOS — {nombre_negocio}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"{fecha_ini.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["Fecha", "Categoría", "Descripción", "Beneficiario", "Monto ($)", "Pago"]
    ws.append(headers)
    header_row = ws.max_row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total = 0.0
    for _, row in df_gastos.iterrows():
        monto = float(row.get('monto', 0))
        total += monto
        ws.append([
            str(row.get('fecha', '')), row.get('categoria', ''),
            row.get('descripcion', ''), row.get('beneficiario', '') or '',
            monto, row.get('tipo_pago', ''),
        ])
        for col_num in range(1, 7):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=4).value = "TOTAL"
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=5).value = total
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


st.title("Gastos")
st.markdown(f"Egresos operativos de: **{nombre_negocio}**")
st.caption(
    "Registra aquí todo el dinero que sale del negocio por gastos operativos: arriendo, servicios, "
    "nómina, transporte, mantenimiento, etc. **La compra de mercancía para vender NO va aquí** — "
    "esa se registra en Inventario → Entradas, para no restar ese costo dos veces de tu utilidad."
)
st.markdown("---")

tab_nuevo, tab_historial, tab_categorias = st.tabs([
    "Registrar Gasto",
    "Historial y Filtros",
    "Resumen por Categoría",
])

# ==========================================
# TAB 1: REGISTRAR GASTO
# ==========================================
with tab_nuevo:
    st.subheader("Registrar Nuevo Gasto")

    with st.form("form_nuevo_gasto", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            fecha_gasto = st.date_input("Fecha del gasto", value=hoy_bogota())
            categoria_gasto = st.selectbox("Categoría", options=CATEGORIAS_GASTO)
            monto_gasto = st.number_input("Monto ($) *", min_value=0.0, step=1000.0)
        with col2:
            beneficiario_gasto = st.text_input("Pagado a (opcional)", placeholder="Ej: EPM, Arrendador, Juan Pérez")
            tipo_pago_gasto = st.selectbox("Forma de pago", options=TIPOS_PAGO_GASTO)
            comprobante_gasto = st.file_uploader(
                "Comprobante (opcional)", type=["png", "jpg", "jpeg", "pdf"],
                help="Foto o PDF del recibo/factura del gasto. Máximo 5MB."
            )

        descripcion_gasto = st.text_input("Descripción *", placeholder="Ej: Arriendo local agosto")
        notas_gasto = st.text_area("Notas (opcional)", height=80)

        st.markdown("")
        if st.form_submit_button("Registrar Gasto", type="primary"):
            if not descripcion_gasto.strip():
                st.warning("La descripción es obligatoria.")
            elif monto_gasto <= 0:
                st.warning("El monto debe ser mayor a cero.")
            elif comprobante_gasto and comprobante_gasto.size > 5 * 1024 * 1024:
                st.error("El comprobante es muy grande. Máximo 5MB.")
            else:
                try:
                    ruta_comprobante = guardar_comprobante(comprobante_gasto, user_id)
                    crear_gasto(
                        uid=user_id,
                        fecha=fecha_gasto,
                        categoria=categoria_gasto,
                        descripcion=descripcion_gasto.strip(),
                        beneficiario=beneficiario_gasto.strip() if beneficiario_gasto else None,
                        monto=monto_gasto,
                        tipo_pago=tipo_pago_gasto,
                        comprobante_path=ruta_comprobante,
                        notas=notas_gasto.strip() if notas_gasto else None,
                    )
                    st.success(f"Gasto de {formato_cop(monto_gasto)} registrado en '{categoria_gasto}'.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

# ==========================================
# TAB 2: HISTORIAL Y FILTROS
# ==========================================
with tab_historial:
    st.subheader("Historial de Gastos")

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        hoy = hoy_bogota()
        hace_30 = hoy - timedelta(days=30)
        fechas_gastos = st.date_input("Rango de fechas", [hace_30, hoy], key="fechas_hist_gastos")
    with col_f2:
        cat_filtro = st.selectbox(
            "Filtrar por categoría", options=["-- Todas --"] + CATEGORIAS_GASTO, key="cat_filtro_gastos"
        )
    with col_f3:
        pago_filtro = st.selectbox(
            "Filtrar por forma de pago", options=["-- Todas --"] + TIPOS_PAGO_GASTO, key="pago_filtro_gastos"
        )

    if len(fechas_gastos) == 2:
        f_ini_g, f_fin_g = fechas_gastos
        df_gastos = obtener_gastos_periodo(user_id, f_ini_g, f_fin_g)

        if cat_filtro != "-- Todas --":
            df_gastos = df_gastos[df_gastos['categoria'] == cat_filtro]
        if pago_filtro != "-- Todas --":
            df_gastos = df_gastos[df_gastos['tipo_pago'] == pago_filtro]

        if not df_gastos.empty:
            total_periodo = float(df_gastos['monto'].sum())
            total_efectivo_periodo = float(df_gastos.loc[df_gastos['tipo_pago'] == 'Efectivo', 'monto'].sum())
            total_transfer_periodo = float(df_gastos.loc[df_gastos['tipo_pago'] == 'Transferencia', 'monto'].sum())

            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Gastos registrados", len(df_gastos))
            col_m2.metric("Total del período", formato_cop(total_periodo))
            col_m3.metric("En efectivo", formato_cop(total_efectivo_periodo))
            col_m4.metric("Por transferencia", formato_cop(total_transfer_periodo))

            st.markdown("---")
            st.dataframe(
                df_gastos[['id', 'fecha', 'categoria', 'descripcion', 'beneficiario', 'monto', 'tipo_pago']].rename(columns={
                    'id': 'N°', 'fecha': 'Fecha', 'categoria': 'Categoría',
                    'descripcion': 'Descripción', 'beneficiario': 'Pagado a',
                    'monto': 'Monto ($)', 'tipo_pago': 'Pago',
                }),
                use_container_width=True, hide_index=True,
                column_config={"Monto ($)": st.column_config.NumberColumn(format="$%,d")},
            )

            excel_gastos = generar_excel_gastos(df_gastos, f_ini_g, f_fin_g, nombre_negocio)
            st.download_button(
                label="Descargar Gastos en Excel",
                data=excel_gastos,
                file_name=f"Gastos_{f_ini_g}_{f_fin_g}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.markdown("---")
            st.markdown("**Editar o eliminar un gasto:**")
            dict_gastos = {
                f"#{r['id']} — {r['fecha']} — {r['categoria']} — {formato_cop(r['monto'])}": r['id']
                for _, r in df_gastos.iterrows()
            }
            gasto_sel_str = st.selectbox(
                "Selecciona el gasto", options=list(dict_gastos.keys()),
                index=None, placeholder="Selecciona un gasto...", key="gasto_sel_editar"
            )

            if gasto_sel_str:
                gasto_id_sel = dict_gastos[gasto_sel_str]
                gasto_actual = obtener_gasto(user_id, gasto_id_sel)

                if gasto_actual:
                    with st.expander("Ver / editar comprobante y notas", expanded=False):
                        if gasto_actual.comprobante_path and os.path.exists(gasto_actual.comprobante_path):
                            if gasto_actual.comprobante_path.lower().endswith((".png", ".jpg", ".jpeg")):
                                st.image(gasto_actual.comprobante_path, width=200)
                            else:
                                with open(gasto_actual.comprobante_path, "rb") as f_comp:
                                    st.download_button(
                                        "Descargar comprobante", data=f_comp.read(),
                                        file_name=os.path.basename(gasto_actual.comprobante_path)
                                    )
                        else:
                            st.caption("Sin comprobante adjunto.")
                        if gasto_actual.notas:
                            st.caption(f"Notas: {gasto_actual.notas}")

                    with st.form("form_editar_gasto"):
                        col_e1, col_e2 = st.columns(2)
                        with col_e1:
                            fecha_edit = st.date_input("Fecha", value=pd.to_datetime(gasto_actual.fecha).date())
                            categoria_edit = st.selectbox(
                                "Categoría", options=CATEGORIAS_GASTO,
                                index=CATEGORIAS_GASTO.index(gasto_actual.categoria)
                                if gasto_actual.categoria in CATEGORIAS_GASTO else len(CATEGORIAS_GASTO) - 1
                            )
                            monto_edit = st.number_input("Monto ($)", min_value=0.0, step=1000.0, value=float(gasto_actual.monto))
                        with col_e2:
                            beneficiario_edit = st.text_input("Pagado a", value=gasto_actual.beneficiario or "")
                            tipo_pago_edit = st.selectbox(
                                "Forma de pago", options=TIPOS_PAGO_GASTO,
                                index=TIPOS_PAGO_GASTO.index(gasto_actual.tipo_pago)
                                if gasto_actual.tipo_pago in TIPOS_PAGO_GASTO else 0
                            )
                            comprobante_edit = st.file_uploader(
                                "Reemplazar comprobante (opcional)", type=["png", "jpg", "jpeg", "pdf"],
                                key="comprobante_edit_uploader"
                            )
                        descripcion_edit = st.text_input("Descripción", value=gasto_actual.descripcion)
                        notas_edit = st.text_area("Notas", value=gasto_actual.notas or "", height=80)

                        col_be1, col_be2 = st.columns(2)
                        guardar_edit = col_be1.form_submit_button("Guardar cambios", type="primary", use_container_width=True)
                        eliminar_edit = col_be2.form_submit_button("Eliminar gasto", use_container_width=True)

                        if guardar_edit:
                            if not descripcion_edit.strip():
                                st.warning("La descripción es obligatoria.")
                            elif monto_edit <= 0:
                                st.warning("El monto debe ser mayor a cero.")
                            else:
                                try:
                                    ruta_nueva = guardar_comprobante(comprobante_edit, user_id)
                                    actualizar_gasto(
                                        uid=user_id, gasto_id=gasto_id_sel, fecha=fecha_edit,
                                        categoria=categoria_edit, descripcion=descripcion_edit.strip(),
                                        beneficiario=beneficiario_edit.strip() if beneficiario_edit else None,
                                        monto=monto_edit, tipo_pago=tipo_pago_edit,
                                        comprobante_path=ruta_nueva, notas=notas_edit.strip() if notas_edit else None,
                                    )
                                    st.success("Gasto actualizado.")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error al actualizar: {e}")

                        if eliminar_edit:
                            try:
                                eliminar_gasto(user_id, gasto_id_sel)
                                st.success("Gasto eliminado.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al eliminar: {e}")
        else:
            st.info("No hay gastos registrados en este período con esos filtros.")
    else:
        st.info("Selecciona un rango de fechas válido.")

# ==========================================
# TAB 3: RESUMEN POR CATEGORÍA
# ==========================================
with tab_categorias:
    st.subheader("¿En qué se va la plata?")

    col_c1, col_c2 = st.columns(2)
    with col_c1:
        hoy_cat = hoy_bogota()
        inicio_mes = hoy_cat.replace(day=1)
        fechas_categoria = st.date_input("Período", [inicio_mes, hoy_cat], key="fechas_categoria_gastos")

    if len(fechas_categoria) == 2:
        f_ini_c, f_fin_c = fechas_categoria
        df_cat = obtener_gastos_por_categoria(user_id, f_ini_c, f_fin_c)

        if not df_cat.empty and float(df_cat['total'].sum()) > 0:
            total_cat = float(df_cat['total'].sum())
            st.metric("Total gastado en el período", formato_cop(total_cat))

            st.bar_chart(df_cat.set_index('categoria')['total'], height=350)

            st.markdown("---")
            df_cat_mostrar = df_cat.copy()
            df_cat_mostrar['porcentaje'] = (df_cat_mostrar['total'] / total_cat * 100).round(1)
            st.dataframe(
                df_cat_mostrar.rename(columns={
                    'categoria': 'Categoría', 'total': 'Total ($)',
                    'cantidad': 'N° de gastos', 'porcentaje': '% del total',
                }),
                use_container_width=True, hide_index=True,
                column_config={
                    "Total ($)": st.column_config.NumberColumn(format="$%,d"),
                    "% del total": st.column_config.NumberColumn(format="%.1f%%"),
                }
            )
        else:
            st.info("No hay gastos registrados en este período.")
